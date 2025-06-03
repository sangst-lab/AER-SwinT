import nibabel as nib
import pydicom as dicom

import os
from torch.utils.data import Dataset
import numpy as np
import albumentations as albu
import matplotlib.pyplot as plt
from openpyxl import load_workbook
import torch

# Custom dataset for gastric carcinoma multi-modal data
class GastricCarcinoma_Dataset(Dataset):
    CLASSES = ['background', 'tumor']

    def __init__(
            self,
            images_masks_dir,
            train_val_test="train",
            need_HU_clip=True,
            augmentation=None,
            preprocessing=None,
    ):
        self.train_val_test = train_val_test
        self.need_HU_clip=need_HU_clip

        # Load file paths and labels for train/val/test
        if self.train_val_test == "train":
            self.images_fps=[]
            self.masks_fps=[]
            self.alllabels=[]
            self.tabular_data=[]

            workbook = load_workbook(images_masks_dir+"/train_label.xlsx")
            sheets = workbook.sheetnames
            label_data = workbook[sheets[0]]
            location_names = {}
            for i, row in enumerate(label_data.iter_rows()):
                if i == 0:
                    continue
                if i == 1:
                    line = [col.value for col in row]
                    for j in range(len(line)):
                        location_names[len(location_names)] = line[j]
                else:
                    line = [col.value for col in row]
                    pat_id = line[0]
                    gender=float(line[1])
                    age=float(line[2]/100)
                    CEA=float(line[4])
                    CA199=float(line[5])
                    location=float(line[6])
                    differentiation=float(line[7])
                    cT=float(line[8])
                    cN=float(line[9])
                    main_task = float(line[10])
                    sub_task_1 = float(line[11])
                    sub_task_2 = float(line[12])

                    # Store CT and mask file paths, and labels
                    imgs=os.listdir(images_masks_dir+"train/"+str(pat_id))
                    for img in imgs:
                        if "IMG" in img:
                            self.images_fps.append(os.path.join(images_masks_dir+"train/" + str(pat_id), img))
                        if "Untitled" in img:
                            self.masks_fps.append(os.path.join(images_masks_dir+"train/" + str(pat_id), img))
                    self.alllabels.append((main_task,sub_task_1,sub_task_2))
                    self.tabular_data.append((gender, age, CEA, CA199, location, differentiation, cT, cN))

        elif self.train_val_test == "val":
            self.images_fps=[]
            self.masks_fps=[]
            self.alllabels=[]
            self.tabular_data=[]

            workbook = load_workbook(images_masks_dir+"/val_label.xlsx")
            sheets = workbook.sheetnames
            label_data = workbook[sheets[0]]
            location_names = {}
            for i, row in enumerate(label_data.iter_rows()):
                if i == 0:
                    continue
                if i == 1:
                    line = [col.value for col in row]
                    for j in range(len(line)):
                        location_names[len(location_names)] = line[j]
                else:
                    line = [col.value for col in row]
                    pat_id = line[0]
                    gender=float(line[1])
                    age=float(line[2]/100)
                    CEA=float(line[4])
                    CA199=float(line[5])
                    location=float(line[6])
                    differentiation=float(line[7])
                    cT=float(line[8])
                    cN=float(line[9])
                    main_task = float(line[10])
                    sub_task_1 = float(line[11])
                    sub_task_2 = float(line[12])

                    # Store CT and mask file paths, and labels
                    imgs=os.listdir(images_masks_dir+"val/"+str(pat_id))
                    for img in imgs:
                        if "IMG" in img:
                            self.images_fps.append(os.path.join(images_masks_dir+"val/" + str(pat_id), img))
                        if "Untitled" in img:
                            self.masks_fps.append(os.path.join(images_masks_dir+"val/" + str(pat_id), img))
                    self.alllabels.append((main_task,sub_task_1,sub_task_2))
                    self.tabular_data.append((gender, age, CEA, CA199, location, differentiation, cT, cN))
        elif self.train_val_test == "test":
            self.images_fps=[]
            self.masks_fps=[]
            self.alllabels=[]
            self.tabular_data=[]
            workbook = load_workbook(images_masks_dir+"/test_label.xlsx")
            sheets = workbook.sheetnames
            label_data = workbook[sheets[0]]
            location_names = {}
            for i, row in enumerate(label_data.iter_rows()):
                if i == 0:
                    continue
                if i == 1:
                    line = [col.value for col in row]
                    for j in range(len(line)):
                        location_names[len(location_names)] = line[j]
                else:
                    line = [col.value for col in row]
                    pat_id = line[0]
                    gender=float(line[1])
                    age=float(line[2]/100)
                    CEA=float(line[4])
                    CA199=float(line[5])
                    location=float(line[6])
                    differentiation=float(line[7])
                    cT=float(line[8])
                    cN=float(line[9])
                    main_task = float(line[10])
                    sub_task_1 = float(line[11])
                    sub_task_2 = float(line[12])
                    # Store CT and mask file paths, and labels
                    imgs=os.listdir(images_masks_dir+"test/"+str(pat_id))
                    for img in imgs:
                        if ".nii" not in img:
                            self.images_fps.append(os.path.join(images_masks_dir+"test/" + str(pat_id), img))
                        if ".nii" in img:
                            self.masks_fps.append(os.path.join(images_masks_dir+"test/" + str(pat_id), img))
                    self.alllabels.append((main_task,sub_task_1,sub_task_2))
                    self.tabular_data.append((gender, age, CEA, CA199, location, differentiation, cT, cN))

        self.augmentation = augmentation
        self.preprocessing = preprocessing

    def __getitem__(self, i):
        # 1. Read DICOM CT image
        organ_dicom = dicom.dcmread(self.images_fps[i])
        intercept = organ_dicom.RescaleIntercept
        slope = organ_dicom.RescaleSlope
        if self.need_HU_clip == True:
            organ_HU = self.HU_clip_min_max((organ_dicom.pixel_array * slope + intercept).astype(np.float32))
        else:
            organ_HU = (organ_dicom.pixel_array * slope + intercept).astype(np.float32)

        # 2. Read NIFTI mask (0: normal, 1: tumor)
        mask_gastcar = np.rot90(np.fliplr(np.squeeze(np.array(nib.load(self.masks_fps[i]).dataobj).astype(np.float32))))

        # Check for different scan coordinate descriptions
        affine = nib.load(self.masks_fps[i]).header.get_best_affine()
        if affine[0, 0] > 0:
            print("Alert: found different affine 0")
        if affine[1, 1] > 0:
            print("Alert: found different affine 1")

        # Crop image and mask to 224x224 around mask center
        w_left, w_right, h_up, h_down=self.crop_points(torch.from_numpy(mask_gastcar))
        mask_gastcar=mask_gastcar[h_up:h_down,w_left:w_right]
        organ_HU=organ_HU[h_up:h_down,w_left:w_right]

        # 3. Only keep tumor region in CT
        RoI=np.multiply(organ_HU,mask_gastcar)

        # Apply augmentation if specified
        if self.augmentation:
            organ_HU = self.augmentation(image=organ_HU,mask=mask_gastcar)
            organ_HU, mask_gastcar = organ_HU['image'], organ_HU['mask']

        # Expand dimensions for model input
        organ_HU=self.dimension_expand(organ_HU)
        mask_gastcar=self.dimension_expand(mask_gastcar)
        RoI=self.dimension_expand(RoI)
        tabular_data=torch.tensor(self.tabular_data[i])

        return organ_HU, mask_gastcar, tabular_data, self.alllabels[i], RoI

    # Find crop region centered on tumor mask
    def crop_points(self,mask, crop_size=(224,224)):
        crop_w = crop_size[0]//2
        crop_h = crop_size[1]//2
        mask=torch.squeeze(mask)
        mask_h=torch.sum(mask,dim=1)
        h1, h2 = np.where(mask_h != 0)[0][0], np.where(mask_h != 0)[0][-1]
        mask_w=torch.sum(mask,dim=0)
        w1, w2 = np.where(mask_w != 0)[0][0], np.where(mask_w != 0)[0][-1]
        h_center=(h2+h1)//2
        w_center=(w2+w1)//2
        return w_center-crop_w, w_center+crop_w, h_center-crop_h, h_center+crop_h

    def __len__(self):
        return len(self.images_fps)

    # Clip and normalize CT values
    def HU_clip_min_max(self, image, min_HU=0, max_HU=300.0):
        np_img = image
        np_img = np.clip(np_img, min_HU, max_HU).astype(np.float32)
        return np_img

    # Expand numpy array to add channel dimension
    def dimension_expand(self,x):
        x=torch.from_numpy(x)
        return torch.unsqueeze(x,0)

# Utility: count folders and find the one with most files
# Returns (number of folders, folder with most files)
def count_folders_and_most_files(directory):
    folder_count = 0
    max_files = 0
    folder_with_most_files = ""
    for root, dirs, files in os.walk(directory):
        if root == directory:
            folder_count = len(dirs)
        if "Untitled.nii.gz" in files:
            current_folder_files = len(files)
            if current_folder_files > max_files:
                max_files = current_folder_files
                folder_with_most_files = root
    return folder_count, folder_with_most_files

# Data augmentation transforms for training/validation
class GastricCarcinomaTransform:
    def get_training_augmentation(self):
        train_transform = [
            # For organs, do not flip as location is important
        ]
        return albu.Compose(train_transform)

    def get_validation_augmentation(self):
        test_transform = [
            albu.PadIfNeeded(224, 224),
        ]
        return albu.Compose(test_transform)

if __name__ == "__main__":
    # same image with different random transforms
    import torch
    import sys
    #sys.path.append("..")
    from visualize import *
    ImgTrans = GastricCarcinomaTransform()
    train_dataset = GastricCarcinoma_Dataset_others_zhongzhongdierpi(
        "D:/data/GastricCarcinoma/other testing data/",
        tabular_data_file="中肿第二批 CT 临床特征 2024-2-19",
        augmentation=ImgTrans.get_validation_augmentation()
    )
    print("----data size----")
    print(len(train_dataset))
    for i in range(0,10):
        img, mask, tablu_data, labels, IoR = train_dataset[i]
        print("----加载图像----")
        print(img.shape)
        print(mask.shape)
        data = train_dataset[i]
        #print(tablu_data)
        img=img.numpy().squeeze()
        mask=mask.numpy().squeeze()
        IoR=IoR.numpy().squeeze()
        #print(img.shape)
        if torch.sum(torch.from_numpy(mask)) > 10:

            figure, axarr = plt.subplots(2, 2)
            axarr[0, 0].imshow(img, cmap='gray')
            axarr[0, 0].axis('off')
            axarr[0, 1].imshow(mask, cmap='gray')
            axarr[0, 1].axis('off')
            axarr[1, 0].imshow(IoR, cmap='gray')
            axarr[1, 0].axis('off')

            #axarr[1, 1].imshow(img, cmap='gray')
            #axarr[1, 1].imshow(mask, cmap='gray', alpha=0.2)
            #axarr[1, 1].axis('off')

            plt.show()
